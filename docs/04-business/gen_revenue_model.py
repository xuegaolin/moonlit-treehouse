# -*- coding: utf-8 -*-
"""
今夜树屋 Moonlit Treehouse · 收入模型 Excel 生成器
用法: python gen_revenue_model.py
产出: revenue-model.xlsx (与本脚本同目录)

设计原则:
- 所有假设集中在「01-核心假设」及各表黄色单元格, 均为可修改的活单元格
- 其余数字全部用公式引用 (=B5*C5 / 跨表 ='01-核心假设'!B5), 改假设即全表联动
- 定价 100% 来自 01-PRD.md, 未自造价格
"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
OUT = os.path.join(HERE, "revenue-model.xlsx")

# ---------------- 样式 ----------------
NIGHT   = "2E2A4F"   # 深夜紫 - 表头
SEC     = "D9D2F0"   # 浅紫 - 分区标题
EDIT    = "FFF3CD"   # 淡黄 - 可调假设
SUB     = "E8E4F5"   # 小计
TOTAL   = "C9BFF0"   # 总计
GOOD    = "E2EFDA"   # 结论绿

F_TITLE = Font(name="微软雅黑", size=14, bold=True, color="FFFFFF")
F_HDR   = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
F_SEC   = Font(name="微软雅黑", size=10, bold=True, color="2E2A4F")
F_BASE  = Font(name="微软雅黑", size=10)
F_BOLD  = Font(name="微软雅黑", size=10, bold=True)
F_NOTE  = Font(name="微软雅黑", size=9, color="808080", italic=True)

FILL_TITLE = PatternFill("solid", fgColor=NIGHT)
FILL_HDR   = PatternFill("solid", fgColor=NIGHT)
FILL_SEC   = PatternFill("solid", fgColor=SEC)
FILL_EDIT  = PatternFill("solid", fgColor=EDIT)
FILL_SUB   = PatternFill("solid", fgColor=SUB)
FILL_TOTAL = PatternFill("solid", fgColor=TOTAL)
FILL_GOOD  = PatternFill("solid", fgColor=GOOD)

_THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT   = Alignment(horizontal="left",  vertical="center", wrap_text=True)
RIGHT  = Alignment(horizontal="right", vertical="center")

FMT_MONEY = "#,##0"
FMT_MONEY1 = "#,##0.0"
FMT_PCT   = "0.0%"
FMT_PCT2  = "0.00%"
FMT_NUM   = "#,##0"

# ---------------- 工具 ----------------
def put(ws, addr, value, font=F_BASE, fill=None, fmt=None, align=None, border=True):
    c = ws[addr]
    c.value = value
    c.font = font
    if fill: c.fill = fill
    if fmt:  c.number_format = fmt
    if align: c.alignment = align
    if border: c.border = BORDER
    return c

def title(ws, addr, text, span):
    ws.merge_cells(f"{addr}:{get_column_letter(ws[addr].column + span - 1)}{ws[addr].row}")
    put(ws, addr, text, font=F_TITLE, fill=FILL_TITLE, align=LEFT)
    ws.row_dimensions[ws[addr].row].height = 26

def section(ws, addr, text, span):
    ws.merge_cells(f"{addr}:{get_column_letter(ws[addr].column + span - 1)}{ws[addr].row}")
    put(ws, addr, text, font=F_SEC, fill=FILL_SEC, align=LEFT)

def note(ws, addr, text, span):
    ws.merge_cells(f"{addr}:{get_column_letter(ws[addr].column + span - 1)}{ws[addr].row}")
    put(ws, addr, text, font=F_NOTE, align=LEFT, border=False)

def widths(ws, wmap):
    for col, w in wmap.items():
        ws.column_dimensions[col].width = w

# 跨表引用常量
S1 = "'01-核心假设'"
S2 = "'02-模块收入模型'"
S3 = "'03-交叉转化敏感度'"
DAYS   = f"{S1}!$B$4"     # 每月天数
DAU0   = f"{S1}!$B$5"     # 基准 DAU
PAY0   = f"{S1}!$B$6"     # 目标付费率 3%
M_CARD = f"{S1}!$B$8"     # 月卡 19
Y_CARD = f"{S1}!$B$9"     # 年卡 128
M_ARPU = f"{S1}!$B$10"    # 会员月均摊销 ARPU
M_MIX  = f"{S1}!$B$11"    # 月卡占比
M_CVR  = f"{S1}!$B$29"    # 会员转化率目标 1.5%
# 模块渗透率 (01 表 C22:C26)
PEN = {m: f"{S1}!$C${22+i}" for i, m in enumerate("ABCDE")}

wb = Workbook()

# ============================================================
# Sheet 01 核心假设
# ============================================================
ws = wb.active
ws.title = "01-核心假设"
ws.sheet_properties.tabColor = NIGHT
title(ws, "A1", "今夜树屋 · 收入模型 — 核心假设(黄色单元格均可修改, 全簿联动)", 6)
note(ws, "A2", "所有定价来自 PRD v1.0; 渗透率/付费率为运营假设, 可随真实数据回测校准。", 6)

section(ws, "A3", "一、全局参数", 6)
rows = [
    ("A4", "每月天数(天)", "B4", 30, FMT_NUM, "按自然月计"),
    ("A5", "基准 DAU(模型计算档)", "B5", 5000, FMT_NUM, "PRD 3 个月目标 DAU 5000"),
    ("A6", "目标付费率(对标 PRD)", "B6", 0.03, FMT_PCT, "PRD 成功指标 ≥3%"),
    ("A7", "目标 ARPU(元/月, 对标 PRD)", "B7", 8, FMT_NUM, "PRD 成功指标 ≥8 元"),
    ("A8", "会员月卡价格(元)", "B8", 19, FMT_NUM, "PRD 定价 19 元/月"),
    ("A9", "会员年卡价格(元)", "B9", 128, FMT_NUM, "PRD 定价 128 元/年"),
    ("A10", "会员月均摊销 ARPU(元)", "B10", "=B8*B11+B9/12*(1-B11)", FMT_MONEY1, "月卡按 19 计, 年卡按 128/12 摊销"),
    ("A11", "会员中月卡购买占比", "B11", 0.7, FMT_PCT, "假设 70% 买月卡、30% 买年卡"),
]
for la, lt, va, vv, fmt, cm in rows:
    put(ws, la, lt, align=LEFT)
    editable = not (isinstance(vv, str) and vv.startswith("="))
    put(ws, va, vv, fmt=fmt, align=RIGHT, fill=FILL_EDIT if editable else None, font=F_BOLD if editable else F_BASE)
    put(ws, f"C{la[1:]}", cm, font=F_NOTE, align=LEFT, border=False)

section(ws, "A13", "二、DAU 分档假设", 6)
for j, h in enumerate(["档位", "DAU", "说明"]):
    put(ws, f"{get_column_letter(1+j)}14", h, font=F_HDR, fill=FILL_HDR, align=CENTER)
for i, (seg, dau, cm) in enumerate([
    ("保守", 1000, "冷启动后低位运行"), ("低速", 3000, "未达成 PRD 目标"),
    ("目标", 5000, "PRD 3 个月目标"), ("乐观", 10000, "裂变超预期"),
]):
    r = 15 + i
    put(ws, f"A{r}", seg, align=CENTER)
    put(ws, f"B{r}", dau, fmt=FMT_NUM, align=RIGHT, fill=FILL_EDIT)
    put(ws, f"C{r}", cm, font=F_NOTE, align=LEFT, border=False)

section(ws, "A20", "三、模块渗透率假设(DAU 中每日进入该模块的比例)", 6)
for j, h in enumerate(["模块", "名称", "日渗透率", "依据"]):
    put(ws, f"{get_column_letter(1+j)}21", h, font=F_HDR, fill=FILL_HDR, align=CENTER)
mods = [
    ("A", "深夜信箱 📮", 0.25, "写信+回信, 情绪浓度高但频次中等"),
    ("B", "摆烂许可证 🪑", 0.40, "一键领取+分享, 门槛最低的流量入口"),
    ("C", "塔罗盲盒 🔮", 0.20, "每日一张免费抽牌, 强仪式感"),
    ("D", "许愿池 🕯", 0.30, "敲木鱼+签到, 高频轻互动"),
    ("E", "漂流墙 💌", 0.35, "刷墙无成本, 浏览型流量"),
]
for i, (m, nm, p, why) in enumerate(mods):
    r = 22 + i
    put(ws, f"A{r}", m, align=CENTER, font=F_BOLD)
    put(ws, f"B{r}", nm, align=LEFT)
    put(ws, f"C{r}", p, fmt=FMT_PCT, align=RIGHT, fill=FILL_EDIT)
    put(ws, f"D{r}", why, font=F_NOTE, align=LEFT, border=False)

section(ws, "A28", "四、会员与成本假设", 6)
cost_rows = [
    (29, "会员转化率目标(PRD ≥1.5%)", 0.015, FMT_PCT, "付费会员 / DAU"),
    (30, "LLM 单次调用成本(元)", 0.02, "0.000", "DeepSeek 级 API 均价, 含回信/解读/祝福信"),
    (31, "日人均 AI 互动率", 0.40, FMT_PCT, "DAU 中触发 AI 功能的用户占比"),
    (32, "日人均 AI 调用次数(次)", 1.5, "0.0", "触发用户平均调用次数"),
    (33, "内容审核单价(元/条)", 0.0015, "0.0000", "微信内容安全 API 量级"),
    (34, "日人均 UGC 条数(条)", 0.16, "0.00", "约 8% 用户日均发 2 条"),
    (35, "微信支付通道费率", 0.006, FMT_PCT, "微信支付标准费率 0.6%"),
]
for r, lt, vv, fmt, cm in cost_rows:
    put(ws, f"A{r}", lt, align=LEFT)
    put(ws, f"B{r}", vv, fmt=fmt, align=RIGHT, fill=FILL_EDIT)
    put(ws, f"C{r}", cm, font=F_NOTE, align=LEFT, border=False)

section(ws, "A37", "五、服务器成本分档(元/月)", 6)
put(ws, "A38", "DAU 上限", font=F_HDR, fill=FILL_HDR, align=CENTER)
put(ws, "B38", "月成本(元)", font=F_HDR, fill=FILL_HDR, align=CENTER)
for i, (cap, cost) in enumerate([(1000, 300), (3000, 500), (5000, 800), (10000, 1500), (999999, 3000)]):
    r = 39 + i
    put(ws, f"A{r}", cap, fmt=FMT_NUM, align=RIGHT, fill=FILL_EDIT)
    put(ws, f"B{r}", cost, fmt=FMT_MONEY, align=RIGHT, fill=FILL_EDIT)
note(ws, "A44", "图例: 淡黄色 = 可调假设; 其余为公式或说明。", 6)
widths(ws, {"A": 30, "B": 14, "C": 34, "D": 34, "E": 10, "F": 10})
ws.freeze_panes = "A3"

# ============================================================
# Sheet 02 模块收入模型
# ============================================================
ws = wb.create_sheet("02-模块收入模型")
ws.sheet_properties.tabColor = "5B4B8A"
title(ws, "A1", "02 · 模块收入模型 — 单点付费(按基准 DAU 计算)", 8)
note(ws, "A2", "日付费人数 = 基准DAU × 模块渗透率 × 该付费点日付费率; 月流水 = 单价 × 日付费人数 × 每月天数。黄色为可调假设。", 8)
hdrs = ["模块", "付费点", "单价(元)", "日付费率(占模块访客)", "日付费人数", "日流水(元)", "月流水(元)", "定价来源(PRD)"]
for j, h in enumerate(hdrs):
    put(ws, f"{get_column_letter(1+j)}3", h, font=F_HDR, fill=FILL_HDR, align=CENTER)
ws.row_dimensions[3].height = 28

# (模块, 付费点, 单价, 日付费率, 来源)
blocks = [
    ("A", "深夜信箱 📮", "A", [
        ("AI 深度回信(500字含心理分析)", 4.9, 0.020, "4.9 元/封"),
        ("定制人格回信(如爱豆口吻)",     9.9, 0.008, "9.9 元/封"),
    ]),
    ("B", "摆烂许可证 🪑", "B", [
        ("高级许可证模板(企业/古风/赛博/宫廷)", 3.0, 0.015, "3 元/张"),
        ("定制头像+姓名专属证书",              6.9, 0.005, "6.9 元"),
        ("周边·真实印刷明信片(含邮费)",        9.9, 0.002, "9.9 元"),
    ]),
    ("C", "塔罗盲盒 🔮", "C", [
        ("三牌阵(过去/现在/未来)",   9.9, 0.012, "9.9 元"),
        ("深度解读-标准档",         19.9, 0.006, "19.9~68 元区间下限"),
        ("深度解读-旗舰档",         68.0, 0.001, "19.9~68 元区间上限"),
        ("情感专题牌阵-标准档",      29.9, 0.004, "29.9~99 元区间下限"),
        ("情感专题牌阵-豪华档",      99.0, 0.001, "29.9~99 元区间上限"),
        ("年度塔罗报告(生日/元旦)",  99.0, 0.0005, "99 元(年化到日均)"),
    ]),
    ("D", "许愿池 🕯", "D", [
        ("高级木鱼皮肤(金色/水晶/敦煌)", 6.0, 0.008, "6 元"),
        ("结愿仪式(AI 心愿达成祝福信)",  4.9, 0.010, "4.9 元"),
    ]),
    ("E", "漂流墙 💌", "E", [
        ("虚拟礼物-糖果档", 0.6, 0.020, "0.6 元档"),
        ("虚拟礼物-抱抱档", 1.9, 0.008, "1.9 元档"),
        ("虚拟礼物-烛光档", 9.9, 0.002, "9.9 元档"),
    ]),
]

r = 4
SUB_ROWS = {}   # 模块 -> 小计行号
for code, name, pen_key, items in blocks:
    start = r
    for k, (pt, price, rate, src) in enumerate(items):
        put(ws, f"A{r}", f"{code} · {name}" if k == 0 else "", align=LEFT, font=F_BOLD if k == 0 else F_BASE)
        put(ws, f"B{r}", pt, align=LEFT)
        put(ws, f"C{r}", price, fmt=FMT_MONEY1, align=RIGHT)          # 单价: PRD 固定价, 不标黄
        put(ws, f"D{r}", rate, fmt="0.00%", align=RIGHT, fill=FILL_EDIT)
        put(ws, f"E{r}", f"={DAU0}*{PEN[pen_key]}*D{r}", fmt=FMT_NUM, align=RIGHT)
        put(ws, f"F{r}", f"=C{r}*E{r}", fmt=FMT_MONEY, align=RIGHT)
        put(ws, f"G{r}", f"=F{r}*{DAYS}", fmt=FMT_MONEY, align=RIGHT)
        put(ws, f"H{r}", src, font=F_NOTE, align=LEFT)
        r += 1
    end = r - 1
    put(ws, f"A{r}", f"{code} 小计", font=F_BOLD, fill=FILL_SUB, align=LEFT)
    for col in "BC":
        put(ws, f"{col}{r}", "", fill=FILL_SUB)
    put(ws, f"D{r}", "", fill=FILL_SUB)
    put(ws, f"E{r}", f"=SUM(E{start}:E{end})", fmt=FMT_NUM, align=RIGHT, fill=FILL_SUB, font=F_BOLD)
    put(ws, f"F{r}", f"=SUM(F{start}:F{end})", fmt=FMT_MONEY, align=RIGHT, fill=FILL_SUB, font=F_BOLD)
    put(ws, f"G{r}", f"=SUM(G{start}:G{end})", fmt=FMT_MONEY, align=RIGHT, fill=FILL_SUB, font=F_BOLD)
    put(ws, f"H{r}", "", fill=FILL_SUB)
    SUB_ROWS[code] = r
    r += 1

TOT_ROW = r
put(ws, f"A{r}", "单点合计", font=F_BOLD, fill=FILL_TOTAL, align=LEFT)
for col in "BCD":
    put(ws, f"{col}{r}", "", fill=FILL_TOTAL)
sub_list = ",".join(f"E{SUB_ROWS[m]}" for m in "ABCDE")
put(ws, f"E{r}", f"=SUM({sub_list})", fmt=FMT_NUM, align=RIGHT, fill=FILL_TOTAL, font=F_BOLD)
put(ws, f"F{r}", f"=SUM({sub_list.replace('E', 'F')})", fmt=FMT_MONEY, align=RIGHT, fill=FILL_TOTAL, font=F_BOLD)
put(ws, f"G{r}", f"=SUM({sub_list.replace('E', 'G')})", fmt=FMT_MONEY, align=RIGHT, fill=FILL_TOTAL, font=F_BOLD)
put(ws, f"H{r}", "", fill=FILL_TOTAL)
r += 1
ws.merge_cells(f"A{r}:F{r}")
put(ws, f"A{r}", "整体付费率(日付费人数 / 基准DAU, 同一用户多点付费未去重, 为上限)", align=LEFT)
put(ws, f"G{r}", f"=E{TOT_ROW}/{DAU0}", fmt=FMT_PCT, align=RIGHT, font=F_BOLD)
put(ws, f"H{r}", "对标 PRD ≥3%", font=F_NOTE, align=LEFT)
RATE_ROW = r
r += 1
ws.merge_cells(f"A{r}:F{r}")
put(ws, f"A{r}", "单点月 ARPU(单点月流水 / 基准DAU)", align=LEFT)
put(ws, f"G{r}", f"=G{TOT_ROW}/{DAU0}", fmt=FMT_MONEY1, align=RIGHT, font=F_BOLD)
put(ws, f"H{r}", "对标 PRD ≥8 元(含会员后)", font=F_NOTE, align=LEFT)
ARPU_ROW = r

widths(ws, {"A": 20, "B": 36, "C": 10, "D": 13, "E": 12, "F": 12, "G": 13, "H": 24})
ws.freeze_panes = "A4"

G_SUB = {m: f"{S2}!$G${SUB_ROWS[m]}" for m in "ABCDE"}   # 模块月流水
E_SUB = {m: f"{S2}!$E${SUB_ROWS[m]}" for m in "ABCDE"}   # 模块日付费人数
G_TOT = f"{S2}!$G${TOT_ROW}"     # 单点月流水合计
G_ARPU = f"{S2}!$G${ARPU_ROW}"   # 单点月 ARPU

# ============================================================
# Sheet 03 交叉转化敏感度 (重点)
# ============================================================
ws = wb.create_sheet("03-交叉转化敏感度")
ws.sheet_properties.tabColor = "B5651D"
title(ws, "A1", "03 · 交叉转化敏感度分析 — 「情绪超市」核心杠杆", 8)
note(ws, "A2", "逻辑: 用户从来源模块被引流到去向模块后, 按去向模块的付费率与 ARPPU 产生增量收入。增量 = DAU × 来源渗透率 × 交叉转化率 × 去向付费率 × 去向ARPPU × 天数。", 8)

# --- Part 1 模块付费效率 ---
section(ws, "A3", "一、模块付费效率(引用 02 表自动计算)", 8)
for j, h in enumerate(["模块", "日访客(人)", "日付费人数", "付费率", "月流水(元)", "ARPPU(元/付费用户·月)"]):
    put(ws, f"{get_column_letter(1+j)}4", h, font=F_HDR, fill=FILL_HDR, align=CENTER)
ws.row_dimensions[4].height = 28
P1 = {}
for i, m in enumerate("ABCDE"):
    r = 5 + i
    P1[m] = r
    put(ws, f"A{r}", f"{m} · {dict(A='深夜信箱',B='摆烂许可证',C='塔罗盲盒',D='许愿池',E='漂流墙')[m]}", align=LEFT)
    put(ws, f"B{r}", f"={DAU0}*{PEN[m]}", fmt=FMT_NUM, align=RIGHT)
    put(ws, f"C{r}", f"={E_SUB[m]}", fmt=FMT_NUM, align=RIGHT)
    put(ws, f"D{r}", f"=C{r}/B{r}", fmt=FMT_PCT, align=RIGHT)
    put(ws, f"E{r}", f"={G_SUB[m]}", fmt=FMT_MONEY, align=RIGHT)
    put(ws, f"F{r}", f"=E{r}/(C{r}*{DAYS})", fmt=FMT_MONEY1, align=RIGHT)

# --- Part 2 交叉转化矩阵 ---
section(ws, "A11", "二、交叉转化矩阵(每日: 从「行」来源模块进入「列」去向模块的转化率, 黄色可调)", 8)
put(ws, "A12", "来源 ↓ / 去向 →", font=F_HDR, fill=FILL_HDR, align=CENTER)
for j, m in enumerate("ABCDE"):
    put(ws, f"{get_column_letter(2+j)}12", f"{m}", font=F_HDR, fill=FILL_HDR, align=CENTER)
MATRIX = {  # (from, to) -> rate
    ("A","B"):0.03, ("A","C"):0.05, ("A","D"):0.02, ("A","E"):0.08,
    ("B","A"):0.01, ("B","C"):0.02, ("B","D"):0.02, ("B","E"):0.01,
    ("C","A"):0.02, ("C","B"):0.01, ("C","D"):0.04, ("C","E"):0.01,
    ("D","A"):0.01, ("D","B"):0.02, ("D","C"):0.03, ("D","E"):0.01,
    ("E","A"):0.05, ("E","B"):0.01, ("E","C"):0.02, ("E","D"):0.02,
}
MROW = {m: 13+i for i, m in enumerate("ABCDE")}
MCOL = {m: get_column_letter(2+j) for j, m in enumerate("ABCDE")}
for m in "ABCDE":
    r = MROW[m]
    put(ws, f"A{r}", f"{m}", font=F_BOLD, fill=FILL_SUB, align=CENTER)
    for n in "ABCDE":
        if m == n:
            put(ws, f"{MCOL[n]}{r}", "—", align=CENTER, fill=FILL_SUB)
        else:
            put(ws, f"{MCOL[n]}{r}", MATRIX[(m, n)], fmt=FMT_PCT, align=CENTER, fill=FILL_EDIT)
note(ws, "A18", "依据: A→E 最高(写信可直接「漂流」到墙, PRD 原生动线); A→C 次之(深夜 emo 写信后顺手抽牌); C→D(塔罗结果引导许愿); E→A(看墙引发倾诉欲)。", 8)

# --- Part 3 逐对增量测算 ---
section(ws, "A20", "三、交叉转化增量测算(按基准 DAU; 行已按增量从大到小排序 → 最上方即最大杠杆)", 8)
for j, h in enumerate(["路径", "转化率", "日交叉引流(人)", "去向付费率", "去向ARPPU(元)", "日增量付费(人)", "月增量流水(元)"]):
    put(ws, f"{get_column_letter(1+j)}21", h, font=F_HDR, fill=FILL_HDR, align=CENTER)
ws.row_dimensions[21].height = 28

# 用默认值预计算排序(仅决定行序, 单元格仍为公式)
f_mod = {}
for m in "ABCDE":
    payers = sum(it[1] for it in [b[3] for b in blocks if b[0] == m][0]) * 5000 * dict(A=0.25, B=0.40, C=0.20, D=0.30, E=0.35)[m]
    rev = sum(it[1] * (5000 * dict(A=0.25, B=0.40, C=0.20, D=0.30, E=0.35)[m] * it[2]) for it in [b[3] for b in blocks if b[0] == m][0]) * 30
    vis = 5000 * dict(A=0.25, B=0.40, C=0.20, D=0.30, E=0.35)[m]
    f_mod[m] = (payers / vis) * (rev / (payers * 30))   # 付费率 × ARPPU
pairs_sorted = sorted(MATRIX.keys(), key=lambda p: -(dict(A=0.25, B=0.40, C=0.20, D=0.30, E=0.35)[p[0]] * MATRIX[p] * f_mod[p[1]]))

PAIR_START = 22
for i, (fm, to) in enumerate(pairs_sorted):
    r = PAIR_START + i
    put(ws, f"A{r}", f"{fm} → {to}", align=CENTER, font=F_BOLD if i < 3 else F_BASE,
        fill=FILL_GOOD if i < 3 else None)
    put(ws, f"B{r}", f"=${MCOL[to]}${MROW[fm]}", fmt=FMT_PCT, align=CENTER)
    put(ws, f"C{r}", f"={DAU0}*{PEN[fm]}*B{r}", fmt=FMT_NUM, align=RIGHT)
    put(ws, f"D{r}", f"=$D${P1[to]}", fmt=FMT_PCT, align=RIGHT)
    put(ws, f"E{r}", f"=$F${P1[to]}", fmt=FMT_MONEY1, align=RIGHT)
    put(ws, f"F{r}", f"=C{r}*D{r}", fmt="0.0", align=RIGHT)
    put(ws, f"G{r}", f"=F{r}*E{r}*{DAYS}", fmt=FMT_MONEY, align=RIGHT,
        fill=FILL_GOOD if i < 3 else None, font=F_BOLD if i < 3 else F_BASE)
PAIR_END = PAIR_START + len(pairs_sorted) - 1   # 41
r = PAIR_END + 1  # 42
SUM_ROW = r
ws.merge_cells(f"A{r}:E{r}")
put(ws, f"A{r}", "交叉转化月增量合计(基准 DAU)", font=F_BOLD, fill=FILL_TOTAL, align=LEFT)
put(ws, f"F{r}", f"=SUM(F{PAIR_START}:F{PAIR_END})", fmt="0.0", align=RIGHT, fill=FILL_TOTAL, font=F_BOLD)
put(ws, f"G{r}", f"=SUM(G{PAIR_START}:G{PAIR_END})", fmt=FMT_MONEY, align=RIGHT, fill=FILL_TOTAL, font=F_BOLD)
INC_TOT = f"$G${SUM_ROW}"
r += 1
put(ws, f"A{r}", "矩阵平均交叉转化率", align=LEFT)
put(ws, f"B{r}", "=AVERAGE(B13:F17)", fmt=FMT_PCT, align=RIGHT, font=F_BOLD)
AVG_RATE = f"$B${r}"
r += 1
put(ws, f"A{r}", "交叉增量率(增量 / 单点月流水)", align=LEFT)
put(ws, f"B{r}", f"={INC_TOT}/{G_TOT}", fmt=FMT_PCT, align=RIGHT, font=F_BOLD)
INC_RATIO = f"{S3}!$B${r}"
r += 1
put(ws, f"A{r}", "基准 DAU 下「单点+交叉」月收入", align=LEFT)
put(ws, f"B{r}", f"={G_TOT}+{INC_TOT}", fmt=FMT_MONEY, align=RIGHT, font=F_BOLD)
r += 1
ws.merge_cells(f"A{r}:F{r}")
put(ws, f"A{r}", "★ 交叉转化率每 +1pp(全矩阵同步), 月增收(基准 DAU)", font=F_BOLD, align=LEFT, fill=FILL_GOOD)
put(ws, f"G{r}", f"={INC_TOT}/{AVG_RATE}*0.01", fmt=FMT_MONEY, align=RIGHT, fill=FILL_GOOD, font=F_BOLD)
STEP1 = f"$G${r}"

# --- Part 4 二维敏感度 ---
SENS_SEC = r + 2
section(ws, f"A{SENS_SEC}", "四、二维敏感度表: 月总收入(单点+交叉, 元) — 交叉转化率 × DAU", 8)
note(ws, f"A{SENS_SEC+1}", "「交叉转化率」为全矩阵同步缩放系数(如 5% 表示矩阵每项均为 5%); 单点收入随 DAU 线性缩放, 交叉增量随 DAU × 转化率线性缩放。", 8)
HR = SENS_SEC + 2   # 表头行
put(ws, f"A{HR}", "交叉转化率 ＼ DAU", font=F_HDR, fill=FILL_HDR, align=CENTER)
for j in range(4):
    put(ws, f"{get_column_letter(2+j)}{HR}", f"={S1}!$B${15+j}", fmt=FMT_NUM, font=F_HDR, fill=FILL_HDR, align=CENTER)
grad = [0.01, 0.02, 0.03, 0.05, 0.08, 0.10]
for i, g in enumerate(grad):
    rr = HR + 1 + i
    put(ws, f"A{rr}", g, fmt=FMT_PCT, align=CENTER, fill=FILL_EDIT, font=F_BOLD)
    for j in range(4):
        col = get_column_letter(2+j)
        f = (f"=({G_TOT}+{INC_TOT}*($A{rr}/{AVG_RATE}))*({col}${HR}/{DAU0})")
        hl = FILL_GOOD if (g == 0.03 and j == 2) else None
        put(ws, f"{col}{rr}", f, fmt=FMT_MONEY, align=RIGHT, fill=hl)
SENS_END = HR + len(grad)
rr = SENS_END + 2
ws.merge_cells(f"A{rr}:A{rr}")
put(ws, f"A{rr}", "★ 每 +1pp 交叉转化率的月增收(元)", font=F_BOLD, align=LEFT, fill=FILL_GOOD)
for j in range(4):
    col = get_column_letter(2+j)
    put(ws, f"{col}{rr}", f"={STEP1}*({col}${HR}/{DAU0})", fmt=FMT_MONEY, align=RIGHT, fill=FILL_GOOD, font=F_BOLD)

widths(ws, {"A": 24, "B": 13, "C": 14, "D": 12, "E": 15, "F": 14, "G": 16, "H": 10})
ws.freeze_panes = "A3"

# ============================================================
# Sheet 04 会员模型
# ============================================================
ws = wb.create_sheet("04-会员模型")
ws.sheet_properties.tabColor = "8E44AD"
title(ws, "A1", "04 · 会员模型 — 会员 vs 单点付费", 6)

section(ws, "A3", "一、会员价格假设(引用 01 表)", 6)
mrows = [
    ("月卡价格(元)", f"={M_CARD}", FMT_NUM, "PRD: 19 元/月"),
    ("年卡价格(元)", f"={Y_CARD}", FMT_NUM, "PRD: 128 元/年"),
    ("月卡购买占比", f"={M_MIX}", FMT_PCT, "可在 01 表修改"),
    ("会员月均摊销 ARPU(元)", f"={M_ARPU}", FMT_MONEY1, "=19×70% + 128/12×30% ≈ 16.5"),
]
for i, (lt, f, fmt, cm) in enumerate(mrows):
    r = 4 + i
    put(ws, f"A{r}", lt, align=LEFT)
    put(ws, f"B{r}", f, fmt=fmt, align=RIGHT, font=F_BOLD)
    put(ws, f"C{r}", cm, font=F_NOTE, align=LEFT, border=False)

section(ws, "A9", "二、会员月收入敏感度(元/月): 会员转化率 × DAU", 6)
put(ws, "A10", "会员转化率 ＼ DAU", font=F_HDR, fill=FILL_HDR, align=CENTER)
for j in range(4):
    put(ws, f"{get_column_letter(2+j)}10", f"={S1}!$B${15+j}", fmt=FMT_NUM, font=F_HDR, fill=FILL_HDR, align=CENTER)
cvr_grad = [0.005, 0.01, f"={M_CVR}", 0.03]
for i, g in enumerate(cvr_grad):
    r = 11 + i
    put(ws, f"A{r}", g, fmt=FMT_PCT, align=CENTER, fill=FILL_EDIT if not isinstance(g, str) else None, font=F_BOLD)
    for j in range(4):
        col = get_column_letter(2+j)
        hl = FILL_GOOD if (i == 2 and j == 2) else None
        put(ws, f"{col}{r}", f"={col}$10*$A{r}*{M_ARPU}", fmt=FMT_MONEY, align=RIGHT, fill=hl)
note(ws, "A15", "第 3 行 1.5% 引用 01 表 PRD 目标转化率, 改 01 表即联动。", 6)

section(ws, "A17", "三、会员 vs 单点(基准 DAU)", 6)
cmp_rows = [
    ("单点月收入(元)", f"={G_TOT}", FMT_MONEY, "02 表合计"),
    ("会员月收入(元, 1.5% 转化)", f"={DAU0}*{M_CVR}*{M_ARPU}", FMT_MONEY, "= DAU × 转化率 × 会员月ARPU"),
    ("交叉转化月增量(元)", f"={S3}!{INC_TOT}", FMT_MONEY, "03 表合计"),
]
for i, (lt, f, fmt, cm) in enumerate(cmp_rows):
    r = 18 + i
    put(ws, f"A{r}", lt, align=LEFT)
    put(ws, f"B{r}", f, fmt=fmt, align=RIGHT)
    put(ws, f"C{r}", cm, font=F_NOTE, align=LEFT, border=False)
put(ws, "A21", "合计月收入(元)", font=F_BOLD, fill=FILL_TOTAL, align=LEFT)
put(ws, "B21", "=SUM(B18:B20)", fmt=FMT_MONEY, align=RIGHT, fill=FILL_TOTAL, font=F_BOLD)
put(ws, "A22", "会员收入占比", align=LEFT)
put(ws, "B22", "=B19/B21", fmt=FMT_PCT, align=RIGHT)
put(ws, "A23", "付费会员数(人)", align=LEFT)
put(ws, "B23", f"={DAU0}*{M_CVR}", fmt=FMT_NUM, align=RIGHT)
put(ws, "A24", "其中: 月卡收入(元)", align=LEFT)
put(ws, "B24", f"=B23*{M_MIX}*{M_CARD}", fmt=FMT_MONEY, align=RIGHT)
put(ws, "A25", "其中: 年卡收入(月摊销, 元)", align=LEFT)
put(ws, "B25", f"=B23*(1-{M_MIX})*{Y_CARD}/12", fmt=FMT_MONEY, align=RIGHT)

section(ws, "A27", "四、蚕食提示(会员权益含 A 无限回信 / C 免费单张, 会替代部分单点)", 6)
put(ws, "A28", "蚕食率假设(会员原单点消费被替代比例)", align=LEFT)
put(ws, "B28", 0.15, fmt=FMT_PCT, align=RIGHT, fill=FILL_EDIT)
put(ws, "A29", "会员净增收(元) = 会员收入 − 会员数×单点月ARPU×蚕食率", align=LEFT)
put(ws, "B29", f"=B19-B23*{G_ARPU}*B28", fmt=FMT_MONEY, align=RIGHT, font=F_BOLD)
note(ws, "A30", "即便如此, 会员的留存价值(不限量权益锁粘性)通常高于单点损失, 且年卡预收改善现金流。", 6)
widths(ws, {"A": 42, "B": 14, "C": 34, "D": 12, "E": 12, "F": 12})
ws.freeze_panes = "A3"

# ============================================================
# Sheet 05 12个月预测
# ============================================================
ws = wb.create_sheet("05-12个月预测")
ws.sheet_properties.tabColor = "16A085"
title(ws, "A1", "05 · 12 个月收入预测(按 PRD 里程碑节奏)", 12)
note(ws, "A2", "模块节奏: M1 仅 B → M2 加 A、C → M3 加 D → M4 加 E → M6 起会员体系(v2.0)。收入系数 = 当月已上线模块月流水 ÷ 全模块月流水(引用 02 表)。", 12)
hdrs = ["月份", "里程碑", "当月已上线模块", "收入系数", "DAU", "单点月收入(元)", "会员转化率", "会员月收入(元)", "交叉系数", "交叉增量(元)", "总月收入(元)", "累计收入(元)"]
for j, h in enumerate(hdrs):
    put(ws, f"{get_column_letter(1+j)}3", h, font=F_HDR, fill=FILL_HDR, align=CENTER)
ws.row_dimensions[3].height = 26

GA, GB, GC, GD, GE = (f"{S2}!$G${SUB_ROWS[m]}" for m in "ABCDE")
coef_f = {
    "B":   f"={GB}/{G_TOT}",
    "BAC": f"=({GA}+{GB}+{GC})/{G_TOT}",
    "BACD": f"=({GA}+{GB}+{GC}+{GD})/{G_TOT}",
    "ALL": "=1",
}
plan = [
    ("M1",  "MVP(W1-2): 摆烂+首页",            "B",         "B",    1000,  0,     0),
    ("M2",  "v1.1 信箱(W3-6) + v1.2 塔罗(M2)", "A+B+C",     "BAC",  2500,  0,     0.5),
    ("M3",  "v1.3 许愿池",                     "A+B+C+D",   "BACD", 5000,  0,     1),
    ("M4",  "v1.4 漂流墙+UGC审核",             "全部",      "ALL",  6000,  0,     1),
    ("M5",  "全模块精细化运营",                 "全部",      "ALL",  7000,  0,     1),
    ("M6",  "v2.0 会员体系+周边电商",           "全部+会员", "ALL",  8000,  0.005, 1),
    ("M7",  "规模增长",                        "全部+会员", "ALL",  9000,  0.01,  1),
    ("M8",  "规模增长",                        "全部+会员", "ALL",  10000, 0.015, 1),
    ("M9",  "规模增长",                        "全部+会员", "ALL",  10500, 0.015, 1),
    ("M10", "规模增长",                        "全部+会员", "ALL",  11000, 0.015, 1),
    ("M11", "规模增长",                        "全部+会员", "ALL",  11500, 0.015, 1),
    ("M12", "规模增长",                        "全部+会员", "ALL",  12000, 0.015, 1),
]
for i, (mo, mile, mods_, coef, dau, mcvr, xcoef) in enumerate(plan):
    r = 4 + i
    put(ws, f"A{r}", mo, align=CENTER, font=F_BOLD)
    put(ws, f"B{r}", mile, align=LEFT)
    put(ws, f"C{r}", mods_, align=CENTER)
    put(ws, f"D{r}", coef_f[coef], fmt=FMT_PCT, align=RIGHT)
    put(ws, f"E{r}", dau, fmt=FMT_NUM, align=RIGHT, fill=FILL_EDIT)
    put(ws, f"F{r}", f"=D{r}*E{r}/{DAU0}*{G_TOT}", fmt=FMT_MONEY, align=RIGHT)
    put(ws, f"G{r}", mcvr, fmt=FMT_PCT, align=RIGHT, fill=FILL_EDIT)
    put(ws, f"H{r}", f"=E{r}*G{r}*{M_ARPU}", fmt=FMT_MONEY, align=RIGHT)
    put(ws, f"I{r}", xcoef, fmt=FMT_PCT, align=RIGHT, fill=FILL_EDIT)
    put(ws, f"J{r}", f"=F{r}*I{r}*{INC_RATIO}", fmt=FMT_MONEY, align=RIGHT)
    put(ws, f"K{r}", f"=F{r}+H{r}+J{r}", fmt=FMT_MONEY, align=RIGHT, font=F_BOLD)
    put(ws, f"L{r}", f"=K{r}" if i == 0 else f"=L{r-1}+K{r}", fmt=FMT_MONEY, align=RIGHT)
r = 16
ws.merge_cells(f"A{r}:J{r}")
put(ws, f"A{r}", "全年合计(元)", font=F_BOLD, fill=FILL_TOTAL, align=LEFT)
put(ws, f"K{r}", "=SUM(K4:K15)", fmt=FMT_MONEY, align=RIGHT, fill=FILL_TOTAL, font=F_BOLD)
put(ws, f"L{r}", "", fill=FILL_TOTAL)
note(ws, "A18", "DAU 假设: M3 达成 PRD 目标 5000; 会员转化率 M6 起步 0.5% → M8 起 1.5%(PRD 目标), 均为可调假设。", 12)
widths(ws, {"A": 7, "B": 30, "C": 12, "D": 9, "E": 9, "F": 13, "G": 11, "H": 13, "I": 10, "J": 13, "K": 13, "L": 14})
ws.freeze_panes = "A4"

# ============================================================
# Sheet 06 汇总看板
# ============================================================
ws = wb.create_sheet("06-汇总看板")
ws.sheet_properties.tabColor = "C0392B"
title(ws, "A1", "06 · 汇总看板 — 关键结论数字", 8)

section(ws, "A3", "一、三档情景月收入(单点+会员+交叉)", 8)
for j, h in enumerate(["情景", "DAU", "付费率假设", "单点月收入(元)", "会员月收入(元)", "交叉增量(元)", "总月收入(元)", "月ARPU(元)"]):
    put(ws, f"{get_column_letter(1+j)}4", h, font=F_HDR, fill=FILL_HDR, align=CENTER)
scen = [
    ("保守", f"={S1}!$B$16", 0.02),
    ("中性", f"={S1}!$B$17", f"={PAY0}"),
    ("乐观", f"={S1}!$B$18", 0.045),
]
for i, (nm, dauf, pr) in enumerate(scen):
    r = 5 + i
    put(ws, f"A{r}", nm, align=CENTER, font=F_BOLD)
    put(ws, f"B{r}", dauf, fmt=FMT_NUM, align=RIGHT)
    put(ws, f"C{r}", pr, fmt=FMT_PCT, align=RIGHT, fill=FILL_EDIT if not isinstance(pr, str) else None)
    put(ws, f"D{r}", f"={G_TOT}*(B{r}/{DAU0})*(C{r}/{PAY0})", fmt=FMT_MONEY, align=RIGHT)
    put(ws, f"E{r}", f"=B{r}*{M_CVR}*{M_ARPU}", fmt=FMT_MONEY, align=RIGHT)
    put(ws, f"F{r}", f"=D{r}*{INC_RATIO}", fmt=FMT_MONEY, align=RIGHT)
    put(ws, f"G{r}", f"=D{r}+E{r}+F{r}", fmt=FMT_MONEY, align=RIGHT, font=F_BOLD,
        fill=FILL_GOOD if nm == "中性" else None)
    put(ws, f"H{r}", f"=G{r}/B{r}", fmt=FMT_MONEY1, align=RIGHT)

section(ws, "A9", "二、月度成本估算(按中性情景 DAU)", 8)
cost_items = [
    ("LLM API 成本(元)", f"=B6*{S1}!$B$31*{S1}!$B$32*{S1}!$B$30*{DAYS}", "DAU×AI互动率×次数×单价×天数"),
    ("内容审核成本(元)", f"=B6*{S1}!$B$34*{S1}!$B$33*{DAYS}", "DAU×人均UGC条数×审核单价×天数"),
    ("服务器成本(元)", (f"=IF(B6<={S1}!$A$39,{S1}!$B$39,IF(B6<={S1}!$A$40,{S1}!$B$40,"
                      f"IF(B6<={S1}!$A$41,{S1}!$B$41,IF(B6<={S1}!$A$42,{S1}!$B$42,{S1}!$B$43))))"), "按 01 表分档"),
    ("微信支付通道费(元)", f"=G6*{S1}!$B$35", "中性总收入 × 0.6%"),
]
for i, (lt, f, cm) in enumerate(cost_items):
    r = 10 + i
    put(ws, f"A{r}", lt, align=LEFT)
    put(ws, f"B{r}", f, fmt=FMT_MONEY, align=RIGHT)
    put(ws, f"C{r}", cm, font=F_NOTE, align=LEFT, border=False)
put(ws, "A14", "成本合计(元)", font=F_BOLD, fill=FILL_TOTAL, align=LEFT)
put(ws, "B14", "=SUM(B10:B13)", fmt=FMT_MONEY, align=RIGHT, fill=FILL_TOTAL, font=F_BOLD)
put(ws, "A15", "月毛利(中性, 元)", align=LEFT)
put(ws, "B15", "=G6-B14", fmt=FMT_MONEY, align=RIGHT, font=F_BOLD, fill=FILL_GOOD)
put(ws, "A16", "毛利率", align=LEFT)
put(ws, "B16", "=B15/G6", fmt=FMT_PCT, align=RIGHT, font=F_BOLD, fill=FILL_GOOD)

section(ws, "A18", "三、盈亏平衡估算", 8)
be_rows = [
    ("变动成本率(LLM+审核+支付费 / 总收入)", "=(B10+B11+B13)/G6", FMT_PCT),
    ("盈亏平衡月收入(元) = 服务器成本 ÷ (1−变动成本率)", "=B12/(1-B19)", FMT_MONEY),
    ("中性月 ARPU(元)", "=H6", FMT_MONEY1),
    ("盈亏平衡 DAU(人) ≈ 平衡收入 ÷ 月ARPU", "=B20/B21", FMT_NUM),
]
for i, (lt, f, fmt) in enumerate(be_rows):
    r = 19 + i
    put(ws, f"A{r}", lt, align=LEFT)
    put(ws, f"B{r}", f, fmt=fmt, align=RIGHT, font=F_BOLD if i >= 2 else F_BASE)
note(ws, "A23", "注: 此模型毛利极高源于「全 AI 无真人」架构, 主要变动成本仅 LLM 调用; 未计入人力与投放成本。", 8)

section(ws, "A25", "四、关键结论", 8)
concl = [
    "① 中性情景(DAU 5000、付费率 3%、会员转化 1.5%)月总收入约 4.3 万元, 月 ARPU ≈ 8.5 元, 达成 PRD ≥8 元目标。",
    "② 交叉转化为第二增长曲线: 默认矩阵下月增收约占单点收入 13%, 且每 +1pp 交叉率 ≈ 基准 DAU 下月增收 ~1900 元(随 DAU 线性放大)。",
    "③ 最大杠杆是「指向 C 塔罗」的交叉路径(A→C、D→C、B→C): 塔罗 ARPPU(≈23 元)为全模块最高, 引流 1 人顶漂流墙 15 人。",
    "④ 会员收入占比仅 ~3%(1.5% 转化下), 短期是留存工具而非收入主力; 年卡占比提升可改善现金流与摊销 ARPU。",
    "⑤ 盈亏平衡 DAU 仅约百人级(LLM 成本极低), 商业风险不在成本侧, 而在 DAU 增长与付费率兑现。",
]
for i, t in enumerate(concl):
    r = 26 + i
    ws.merge_cells(f"A{r}:H{r}")
    put(ws, f"A{r}", t, align=LEFT, fill=FILL_GOOD if i in (0, 2) else None)
widths(ws, {"A": 44, "B": 13, "C": 26, "D": 14, "E": 14, "F": 13, "G": 14, "H": 12})
ws.freeze_panes = "A3"

wb.save(OUT)
print(f"OK -> {OUT}")
print("Sheets:", wb.sheetnames)
